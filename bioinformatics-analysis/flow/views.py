import json

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import load_workbook

from utils.response import response_body

from common.viewsets.viewsets import CustomeViewSets
from flow.models import Flow, FlowMembers, Flow2Sample, PanelGroup, Panel
from flow.filters import (
    FlowFilters,
    FilterByAccount,
    PanelFilters,
    CustomPanelFilterSet,
)
from flow.serializers import FlowSerializer, PanelGroupSerializer, PanelSerializer
from utils.paginator import PageNumberPaginationWithWrapper


@csrf_exempt
def upload_excel(request):
    """解析上传上来的 excel 文件."""
    if request.method == "POST" and request.FILES.get("file"):
        excel_file = request.FILES["file"]

        wb = load_workbook(excel_file)
        sheet = wb.active

        data = []
        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {headers[i]: row[i] for i in range(len(headers))}
            data.append(row_data)

        return response_body(data=data)

    return response_body(status_code=400, code=-1, msg="request invalid")


class PanelGroupView(CustomeViewSets):
    queryset = PanelGroup.objects.all()
    serializer_class = PanelGroupSerializer

    filter_backends = [FilterByAccount]


class PanelView(CustomeViewSets):
    queryset = Panel.objects.prefetch_related("panel_group").all()
    serializer_class = PanelSerializer
    # pagination_class = PageNumberPaginationWithWrapper

    # filter_backends = [DjangoFilterBackend, FilterByAccount, PanelFilters]
    filter_backends = [DjangoFilterBackend, PanelFilters]
    filterset_class = CustomPanelFilterSet


class FlowView(CustomeViewSets):
    queryset = Flow.objects.all()
    serializer_class = FlowSerializer
    pagination_class = PageNumberPaginationWithWrapper

    # 暂不进行权限过滤
    # filter_backends = [FilterByAccount, FlowFilters]
    filter_backends = [FlowFilters]

    def create_data(self, request, *args, **kwargs):
        data = super().create_data(request, *args, **kwargs)
        data["parameter_schema"] = json.dumps(data.get("parameters", "{}"))
        data["owner_id"] = request.account.id
        return data

    def update_data(self, request, *args, **kwargs):
        not_supported_fields = ["id", "owner_id", "flow_type"]
        data = super().update_data(request, *args, **kwargs)
        data["owner_id"] = request.account.id
        if isinstance(data["parameters"], str):
            data["parameter_schema"] = data["parameters"]
        else:
            data["parameter_schema"] = json.dumps(data["parameters"])
        for f in not_supported_fields:
            if f in data:
                data.pop(f)

        return data

    def list_types(self, request, *args, **kwargs):
        fields = ["flow_category", "alignment_tool"]

        objects = Flow.objects.all()
        if "admin" not in request.role_list:
            account_id = request.account.id
            flow_ids = FlowMembers.objects.filter(
                account_id=account_id).values_list("flow_id")
            objects = objects.filter(id__in=flow_ids)

        data = {}
        for f in fields:
            data[f] = [
                v for v in objects.order_by(f).distinct().values_list(
                    "flow_category", flat=True) if v
            ]

        return response_body(data=data)

    def query(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)

    def add_members(self, request, *args, **kwargs):
        if "admin" not in request.role_list:
            return response_body(code=1, msg="无权限设置")

        payload = json.loads(request.body)
        account_ids = payload.get("account_ids", [])
        flow_ids = payload["flow_ids"]

        for account_id in account_ids:
            FlowMembers.objects.filter(account_id=account_id).delete()

        for flow_id in flow_ids:
            for account_id in account_ids:
                FlowMembers.objects.get_or_create(account_id=account_id,
                                                  flow_id=flow_id)

        return response_body(data=[], msg="success")

    def remove_members(self, request, *args, **kwargs):
        if "admin" not in request.role_list:
            return response_body(code=1, msg="无权限设置")

        payload = json.loads(request.body)
        account_ids = payload.get("account_ids", [])
        flow_ids = payload["flow_ids"]

        FlowMembers.objects.filter(flow_id__in=flow_ids,
                                   account_id__in=account_ids).delete()

        return response_body(data=[], msg="success")

    def list_samples(self, request, *args, **kwargs):
        data = request.GET
        sample_ids = set(
            Flow2Sample.objects.filter(
                project_id=data.get("project_id"),
                flow_id=data.get("flow_id")).values_list("sample_id",
                                                         flat=True))
        return response_body(data=list(sample_ids))
